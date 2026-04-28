"""Extract SAFE constraint indicator from country files.

Handles two formats and correctly distinguishes:
- Q0: "What is currently the most important problem your firm is facing?" -> SHARE (%)
- Q0b: "How important have the following problems been..." -> MEAN SCORE (1-10)

We need the Q0 SHARE, not the Q0b mean score.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook

from src.config import DATA_RAW, DATA_INTERIM

COUNTRY_MAP = {
    "austria": "AT",
    "belgium": "BE",
    "bulgaria": "BG",
    "croatia": "HR",
    "cyprus": "CY",
    "czechia": "CZ",
    "czech republic": "CZ",
    "denmark": "DK",
    "estonia": "EE",
    "finland": "FI",
    "france": "FR",
    "germany": "DE",
    "greece": "EL",
    "hungary": "HU",
    "ireland": "IE",
    "italy": "IT",
    "latvia": "LV",
    "lithuania": "LT",
    "luxembourg": "LU",
    "malta": "MT",
    "netherlands": "NL",
    "poland": "PL",
    "portugal": "PT",
    "romania": "RO",
    "slovakia": "SK",
    "slovenia": "SI",
    "spain": "ES",
    "sweden": "SE",
    "united kingdom": "UK",
}


def name_to_iso(name: str) -> str | None:
    n = name.lower().strip()
    return COUNTRY_MAP.get(n)


def extract_safe(filepath: Path) -> pd.DataFrame | None:
    year = int(filepath.stem.split("_")[-1])
    wb = load_workbook(filepath, data_only=True, read_only=True)

    # Find sheet with Q0 (not Q0b)
    sheet_name = None
    for name in wb.sheetnames:
        ws = wb[name]
        max_rows = 200 if len(wb.sheetnames) == 1 else 20
        for row in ws.iter_rows(max_row=max_rows, values_only=True):
            text = " ".join(str(x).lower() for x in row if x).strip()
            # Match Q0 about "most important problem" but NOT Q0b about "following problems"
            if "most important problem" in text and "following problems" not in text:
                sheet_name = name
                break
        if sheet_name:
            break

    if not sheet_name:
        print(f"  [SKIP] {filepath.name}: no Q0 found")
        return None

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find Q0 question row, then header row below it, then data rows
    q0_idx = None
    for i, row in enumerate(rows):
        text = " ".join(str(x).lower() for x in row if x).strip()
        if "most important problem" in text and "following problems" not in text:
            q0_idx = i
            break

    if q0_idx is None:
        return None

    # Search for header row (with country names) after Q0
    header_idx = None
    for i in range(q0_idx + 1, min(q0_idx + 5, len(rows))):
        text = " ".join(str(x).lower() for x in rows[i] if x).strip()
        if "austria" in text or "belgium" in text:
            header_idx = i
            break

    if header_idx is None:
        print(f"  [SKIP] {filepath.name}: no header after Q0")
        return None

    # Search for "access to finance" data row after header
    data_idx = None
    for i in range(header_idx + 1, min(header_idx + 15, len(rows))):
        text = " ".join(str(x).lower() for x in rows[i] if x).strip()
        if "access to finance" in text:
            data_idx = i
            break

    if data_idx is None:
        print(f"  [SKIP] {filepath.name}: no access to finance row")
        return None

    header = rows[header_idx]
    data = rows[data_idx]

    records = []
    for j, val in enumerate(header):
        if j == 0 or val is None:
            continue
        iso = name_to_iso(str(val))
        if iso and j < len(data):
            try:
                v = float(data[j])
                records.append(
                    {"country": iso, "year": year, "access_to_finance_share": v}
                )
            except (ValueError, TypeError):
                pass

    if not records:
        print(f"  [SKIP] {filepath.name}: no records")
        return None

    print(f"  [OK] {filepath.name}: {len(records)} countries")
    return pd.DataFrame(records)


if __name__ == "__main__":
    print("=== Extracting SAFE constraint panel ===\n")
    files = sorted(DATA_RAW.glob("SAFE_resbycountry_*.xlsx"))
    print(f"Files: {len(files)}")

    dfs = []
    for f in files:
        df = extract_safe(f)
        if df is not None:
            dfs.append(df)

    if not dfs:
        print("No data extracted!")
        sys.exit(1)

    panel = pd.concat(dfs, ignore_index=True)
    print(f"\nPanel: {len(panel)} rows")
    print(f"Years: {sorted(panel['year'].unique())}")
    print(
        f"Countries ({panel['country'].nunique()}): {sorted(panel['country'].unique())}"
    )

    # Filter to EU-27 (drop UK)
    panel = panel[panel["country"] != "UK"]
    print(
        f"After dropping UK: {len(panel)} rows, {panel['country'].nunique()} countries"
    )

    out = DATA_INTERIM / "safe_constraint_panel.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    panel.to_csv(out, index=False)
    print(f"\nSaved to {out}")

    # Summary stats
    print(
        f"\nValue range: {panel['access_to_finance_share'].min():.4f} - {panel['access_to_finance_share'].max():.4f}"
    )
    print(f"Mean: {panel['access_to_finance_share'].mean():.4f}")
