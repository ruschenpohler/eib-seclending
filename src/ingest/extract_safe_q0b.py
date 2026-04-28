"""Extract Q0b mean importance score from SAFE country files.

Q0b: "How important have the following problems been for your enterprise
in the past six months? Please answer on a scale of 1-10"

We extract the mean score for "access to finance".
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook

from src.config import DATA_RAW, DATA_INTERIM

COUNTRY_MAP = {
    "austria": "AT",
    "belgium": "BE",
    "bulgaria": "BG",
    "croatia": "HR",
    "cyprus": "CY",
    "czechia": "CZ",
    "czech republic": "CZ",
    "denmark": "DK",
    "estonia": "EE",
    "finland": "FI",
    "france": "FR",
    "germany": "DE",
    "greece": "EL",
    "hungary": "HU",
    "ireland": "IE",
    "italy": "IT",
    "latvia": "LV",
    "lithuania": "LT",
    "luxembourg": "LU",
    "malta": "MT",
    "netherlands": "NL",
    "poland": "PL",
    "portugal": "PT",
    "romania": "RO",
    "slovakia": "SK",
    "slovenia": "SI",
    "spain": "ES",
    "sweden": "SE",
    "united kingdom": "UK",
}


def name_to_iso(name: str):
    return COUNTRY_MAP.get(name.lower().strip())


def extract_q0b(filepath: Path):
    year = int(filepath.stem.split("_")[-1])
    wb = load_workbook(filepath, data_only=True, read_only=True)

    # For single-sheet files (2015-2020), Q0b is on the main sheet
    # For multi-sheet (2021+), Q0b is in T9 or similar
    sheet_name = wb.sheetnames[0] if len(wb.sheetnames) == 1 else None

    if sheet_name is None:
        # Multi-sheet: search for Q0b
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(max_row=10, values_only=True):
                text = " ".join(str(x).lower() for x in row if x).strip()
                if "q0b" in text or "following problems" in text:
                    sheet_name = name
                    break
            if sheet_name:
                break

    if not sheet_name:
        return None

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find Q0b section
    q0b_idx = None
    for i, row in enumerate(rows):
        text = " ".join(str(x).lower() for x in row if x).strip()
        if "following problems" in text and "scale" in text:
            q0b_idx = i
            break

    if q0b_idx is None:
        return None

    # Find header row after Q0b
    header_idx = None
    for i in range(q0b_idx + 1, min(q0b_idx + 10, len(rows))):
        text = " ".join(str(x).lower() for x in rows[i] if x).strip()
        if "austria" in text or "belgium" in text:
            header_idx = i
            break

    if header_idx is None:
        return None

    # Find "access to finance" data row
    data_idx = None
    for i in range(header_idx + 1, min(header_idx + 15, len(rows))):
        text = " ".join(str(x).lower() for x in rows[i] if x).strip()
        if "access to finance" in text:
            data_idx = i
            break

    if data_idx is None:
        return None

    header = rows[header_idx]
    data = rows[data_idx]

    records = []
    for j, val in enumerate(header):
        if j == 0 or val is None:
            continue
        iso = name_to_iso(str(val))
        if iso and j < len(data):
            try:
                v = float(data[j])
                records.append(
                    {
                        "country": iso,
                        "year": year,
                        "access_to_finance_mean_score": v,
                    }
                )
            except (ValueError, TypeError):
                pass

    if not records:
        return None

    return pd.DataFrame(records)


if __name__ == "__main__":
    files = sorted(DATA_RAW.glob("SAFE_resbycountry_*.xlsx"))
    dfs = [extract_q0b(f) for f in files]
    dfs = [d for d in dfs if d is not None]

    if not dfs:
        print("No Q0b data extracted!")
        sys.exit(1)

    panel = pd.concat(dfs, ignore_index=True)
    panel = panel[panel["country"] != "UK"]

    out = DATA_INTERIM / "safe_q0b_panel.csv"
    panel.to_csv(out, index=False)
    print(f"Saved {out}: {len(panel)} rows, {panel['country'].nunique()} countries")
    print(f"Years: {sorted(panel['year'].unique())}")
